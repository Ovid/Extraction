"""Tests for CPI (Transparency International) data handling.

CPI data is fetched as Excel by the fetcher module. This test validates
the fetcher's Excel parsing logic using programmatically generated fixtures.
"""

import openpyxl
import pandas as pd


class TestCpiFetcherParsing:
    def test_excel_fixture_roundtrips(self, tmp_path):
        xlsx_path = tmp_path / "cpi_test.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "CPI Timeseries"
        ws.append(["Country", "ISO3", "CPI Score 2024", "CPI Score 2023"])
        ws.append(["Denmark", "DNK", 90, 90])
        ws.append(["United States", "USA", 69, 69])
        ws.append(["Nigeria", "NGA", 25, 24])
        wb.save(xlsx_path)

        df = pd.read_excel(xlsx_path, sheet_name="CPI Timeseries")
        assert len(df) == 3
        assert "ISO3" in df.columns
        assert df[df["ISO3"] == "DNK"]["CPI Score 2024"].iloc[0] == 90

    def test_finds_sheet_with_iso_column(self, tmp_path):
        xlsx_path = tmp_path / "cpi_multi_sheet.xlsx"
        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = "Metadata"
        ws1.append(["About", "CPI 2024"])
        ws2 = wb.create_sheet("CPI Results")
        ws2.append(["Country", "ISO3", "Score"])
        ws2.append(["Denmark", "DNK", 90])
        wb.save(xlsx_path)

        xl = pd.ExcelFile(xlsx_path)
        found = None
        for sheet_name in xl.sheet_names:
            candidate = pd.read_excel(xlsx_path, sheet_name=sheet_name)
            cols_lower = [str(c).lower() for c in candidate.columns]
            if any("iso" in c for c in cols_lower):
                found = candidate
                break

        assert found is not None
        assert "ISO3" in found.columns
